"""
parse_spreadsheet.py

This module handles spreadsheet files (Excel .xlsx, .xls, or CSV) by extracting
textual content in a structured way. The primary goal is to return a dictionary
with the following keys:

{
  "text":   <string concatenating all sheet data>,
  "tables": <list of 2D table data, one entry per sheet or per CSV>,
  "images": <list of placeholders or empty, as spreadsheets rarely store images in this approach>,
  "metadata": {
      "file_name": <source spreadsheet file>,
      "sheet_count": <number of sheets or 1 if CSV>,
      ...
  }
}

Guiding Principles:
1. Offline, open-source approach: use openpyxl for Excel, Python's csv or pandas for CSV.
2. Modular design: parse_spreadsheet.py focuses on reading spreadsheets, returning consistent data.
3. Detailed comments for clarity and maintainability.
4. Consistent return structure with parse_pdf.py, parse_docx.py, etc.
   (so the pipeline can handle them uniformly).

Notes:
- openpyxl can read .xlsx (and .xlsm) files reliably. For .xls (older Excel), you might need
  a different approach or library (like xlrd). We'll assume modern .xlsx format or you can
  add checks for .xls.
- For CSV, Python's built-in csv or pandas can be used. We'll demonstrate a simple approach
  with the built-in csv library. If you prefer pandas, you can adapt accordingly.
- Spreadsheets can contain multiple sheets (openpyxl). We'll gather text from each sheet row by row.

Usage:
    from parse_spreadsheet import parse_spreadsheet

    result = parse_spreadsheet("path/to/data.xlsx")
    # result is a dict with "text", "tables", "images", "metadata"
"""

import os
import sys
import csv

# openpyxl for Excel .xlsx
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    # If openpyxl isn't installed, this script won't handle .xlsx gracefully.

def parse_spreadsheet(file_path: str) -> dict:
    """
    Reads an Excel (.xlsx) or CSV file from disk, extracting textual data in
    both a combined text form and as table-like structures. Returns a dictionary
    that fits the standard ingestion pattern:

      {
        "text": "<combined textual content>",
        "tables": [ list-of-rows for each sheet ],
        "images": [],
        "metadata": { "file_name": ..., "sheet_count": ..., ... }
      }

    :param file_path: The path to an Excel or CSV file.
    :type file_path: str

    :return: Dictionary with keys "text", "tables", "images", "metadata".
    :rtype: dict

    Detailed Steps:
      1) Check if file exists, raise FileNotFoundError if not.
      2) Check file extension. If .xlsx (or .xlsm, .xltx, etc.), use openpyxl.
         If .csv, parse with built-in csv. (For .xls or other types, you may need
         another library or logic.)
      3) For Excel:
         - load_workbook
         - iterate over sheets and rows
         - gather text from each cell, create a "table" per sheet
         - store them in 'tables' and also build a big 'text' string
      4) For CSV:
         - open the CSV in read mode, parse rows
         - store them as a single "table" in 'tables'
         - also build a big 'text' string from row data
      5) Return the final dictionary

    Dependencies:
      - openpyxl for .xlsx
      - built-in csv for .csv
    """

    if not os.path.isfile(file_path):
        raise FileNotFoundError(
            f"parse_spreadsheet: File '{file_path}' not found or inaccessible."
        )

    extension = os.path.splitext(file_path)[1].lower()
    if extension in [".xlsx", ".xlsm", ".xltx", ".xltm"]:  # extended Excel
        if not HAS_OPENPYXL:
            raise ImportError(
                "parse_spreadsheet: openpyxl is not installed. "
                "Cannot parse Excel files without openpyxl."
            )
        return _parse_excel(file_path)
    elif extension in [".csv"]:
        return _parse_csv(file_path)
    else:
        # If .xls or other extension, you might raise an error or handle differently
        # For now, we only handle .xlsx or .csv in this script
        raise ValueError(
            f"parse_spreadsheet: Unsupported file extension '{extension}'. "
            "Please provide .xlsx or .csv."
        )


def _parse_excel(file_path: str) -> dict:
    """
    Helper for parse_spreadsheet to handle Excel files (.xlsx, .xlsm, .xltx, .xltm).
    Returns the standard dictionary structure with "text", "tables", etc.
    """
    # Load the workbook in read-only mode if you have large files
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    # We'll store text for each sheet, then combine
    all_sheets_text = []
    # We'll store a table for each sheet in 'tables'
    all_tables = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        # We'll build a 2D list for the current sheet
        current_sheet_table = []
        # We'll also store textual representation row by row
        sheet_text_rows = []

        # iterate over rows. For read_only=True, you can do sheet.iter_rows(...)
        for row in sheet.iter_rows(values_only=True):
            row_data = []
            # row is a tuple of cell values, e.g. (val1, val2, val3,...)
            # Convert each cell to a string for uniformity (or handle types differently if needed)
            for cell_val in row:
                cell_str = str(cell_val) if cell_val is not None else ""
                row_data.append(cell_str)
            current_sheet_table.append(row_data)
            # Also create a textual line for the row
            sheet_text_rows.append(" | ".join(row_data).strip())

        # Add the table data from this sheet to all_tables
        all_tables.append(current_sheet_table)
        # Combine all rows from this sheet into a single text block, preceded by sheet name
        sheet_text = f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_text_rows)
        all_sheets_text.append(sheet_text)

    # Combine text from all sheets, separated by blank lines
    combined_text = "\n\n".join(all_sheets_text)

    metadata = {
        "file_name": os.path.basename(file_path),
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names
    }

    parse_result = {
        "text": combined_text,
        "tables": all_tables,
        "images": [],  # Typically no images in standard Excel approach
        "metadata": metadata
    }

    # Close the workbook to free resources
    wb.close()
    return parse_result


def _parse_csv(file_path: str) -> dict:
    """
    Helper for parse_spreadsheet to handle CSV files. Returns the standard structure.
    Uses Python's built-in csv module for a simple approach (one table).
    """
    all_rows = []
    # We store text line by line
    text_rows = []

    with open(file_path, "r", encoding="utf-8", newline="") as f:
        # Attempt to sniff the dialect
        dialect = None
        try:
            dialect = csv.Sniffer().sniff(f.read(1024))
            f.seek(0)
        except csv.Error:
            # If sniff fails, we'll just use default
            f.seek(0)
            pass

        reader = csv.reader(f, dialect=dialect) if dialect else csv.reader(f)
        for row in reader:
            # row is a list of column strings
            all_rows.append(row)
            text_rows.append(" | ".join(row))

    # We treat CSV as a single sheet with rows
    # Convert 'all_rows' to a single text block
    combined_text = "--- CSV Content ---\n" + "\n".join(text_rows)

    metadata = {
        "file_name": os.path.basename(file_path),
        "sheet_count": 1,
        "sheet_names": ["(csv)"]
    }

    parse_result = {
        "text": combined_text,
        "tables": [all_rows],  # list with one element (the entire CSV as a table)
        "images": [],
        "metadata": metadata
    }

    return parse_result


if __name__ == "__main__":
    """
    CLI usage:
      python parse_spreadsheet.py path/to/file.xlsx
      or
      python parse_spreadsheet.py path/to/data.csv

    It will attempt to parse the file and print some details to stdout.
    """
    if len(sys.argv) < 2:
        print("Usage: python parse_spreadsheet.py <spreadsheet_file>")
        sys.exit(1)

    input_file = sys.argv[1]

    try:
        result = parse_spreadsheet(input_file)
        print(f"Successfully parsed '{input_file}'.\n")

        # Show some text
        print("--- Extracted Text (first 500 chars) ---")
        snippet = result["text"][:500]
        more_text = "..." if len(result["text"]) > 500 else ""
        print(snippet + more_text)
        print("----------------------------------------")

        # Show table info
        print(f"\nNumber of tables: {len(result['tables'])}")
        if result["tables"]:
            print(f"Rows in first table: {len(result['tables'][0])}")
            print("First 5 rows of first table:")
            for row in result["tables"][0][:5]:
                print(row)

        print("\nMetadata:", result["metadata"])
        print("Images extracted:", len(result["images"]))
    except Exception as e:
        print(f"Error parsing spreadsheet '{input_file}': {e}")
