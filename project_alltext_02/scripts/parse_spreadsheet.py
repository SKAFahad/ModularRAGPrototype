"""
parse_spreadsheet.py

A pipeline that:
 - Loads an Excel workbook with openpyxl
 - Iterates over each worksheet
 - Extracts table data (2D list) + text from each cell
 - Extracts embedded images (if any)
 - Returns a dictionary:
    {
      "text":   <SINGLE BIG STRING of all sheets' text>,
      "tables": [list of DataFrames],
      "images": [list or dictionary of images, as you prefer]
    }

We've changed the code so that it no longer returns text as a dict {sheet_name: "..."}.
Instead, we produce one big string combining each sheet's text, ensuring aggregator can
write it directly to a .txt file.
"""

import os
import base64
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image  # for extracting embedded images, if present

def extract_text_and_table(ws):
    """
    Extracts cell values from a worksheet (ws).
    Returns:
      table_data: a 2D list (list of rows), each row is a list of cell strings
      full_text:  a string built by concatenating non-empty cell contents
    """
    table_data = []
    all_text = []

    # Iterate over rows, retrieving only the cell values
    for row in ws.iter_rows(values_only=True):
        row_data = []
        for cell in row:
            cell_value = str(cell) if cell is not None else ""
            row_data.append(cell_value)
            if cell_value.strip():
                all_text.append(cell_value.strip())
        table_data.append(row_data)

    # Join all non-empty cell values with newlines for a single text block
    full_text = "\n".join(all_text)
    return table_data, full_text

def extract_images(ws, output_dir, sheet_name):
    """
    Extracts embedded images from a worksheet's _images attribute.
    Saves them to output_dir, returning the file paths in a list.
    This is for reference if you want to store images. The aggregator
    might do the actual saving instead—adapt to your pipeline needs.
    """
    image_paths = []
    if hasattr(ws, '_images'):
        for idx, img_obj in enumerate(ws._images, start=1):
            ext = ".png"
            if img_obj.path:
                _, ext_guess = os.path.splitext(img_obj.path)
                if ext_guess:
                    ext = ext_guess
            filename = f"{sheet_name}_image_{idx}{ext}"
            filepath = os.path.join(output_dir, filename)
            img_obj.image.save(filepath)
            image_paths.append(filepath)
    return image_paths

def parse_spreadsheet(excel_path, output_dir="excel_output"):
    """
    Main function that returns a dictionary for the aggregator:
      {
        "text":   <string of all sheets' text combined>,
        "tables": [DataFrames for each sheet],
        "images": [],  # or a list of references if you want
      }

    Also demonstrates saving tables as CSV, but aggregator might
    handle that— adapt as needed.
    """
    if not os.path.isfile(excel_path):
        raise ValueError(f"[parse_spreadsheet] File not found: {excel_path}")

    os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(excel_path, data_only=True)

    all_text_fragments = []
    all_tables = []
    all_images = []  # optional, if you want to store references

    for ws in wb.worksheets:
        sheet_name = ws.title
        print(f"[parse_spreadsheet] Processing sheet: {sheet_name}")

        # 1) Extract table data + text
        table_data, sheet_text = extract_text_and_table(ws)
        # Accumulate text into an overall list
        if sheet_text.strip():
            all_text_fragments.append(f"--- Sheet: {sheet_name} ---\n{sheet_text}\n")

        # Convert table_data to a DataFrame
        df = pd.DataFrame(table_data)
        all_tables.append(df)

        # Optionally save table as CSV right here (or let aggregator handle it)
        csv_filename = f"{sheet_name}_table.csv"
        csv_path = os.path.join(output_dir, csv_filename)
        df.to_csv(csv_path, index=False, header=False)
        print(f"[parse_spreadsheet] Saved CSV: {csv_filename}")

        # 2) Extract images from the sheet
        image_files = extract_images(ws, output_dir, sheet_name)
        all_images.extend(image_files)

    # Combine all sheet text into a single big string
    # The aggregator expects "text" to be directly writeable
    unified_text = "\n".join(all_text_fragments)

    # Optionally also store images or references
    # aggregator might handle them differently
    # For now, we'll store a blank list or the file paths
    return {
        "text": unified_text,
        "tables": all_tables,    # list of DataFrames
        "images": all_images     # list of image file paths
    }

if __name__ == "__main__":
    # Example usage
    excel_file = "your_spreadsheet.xlsx"
    output = parse_spreadsheet(excel_file, "excel_output")
    print("parse_spreadsheet returned:")
    print(" text length:", len(output["text"]))
    print(" tables:", len(output["tables"]))
    print(" images:", len(output["images"]))
